"""Tests for dsno_processor.ebs_download (non-Selenium helpers).

Only tests the pure-Python helpers: history tracking and
``read_files_from_excel``.  Browser-dependent functions are excluded.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from dsno_processor.ebs_download import (
    already_downloaded,
    load_history,
    read_files_from_excel,
    record_success,
    save_history,
)
from dsno_processor.exceptions import ConfigurationError


# ── History helpers ──────────────────────────────────────────────────


class TestHistoryHelpers:
    """Tests for load/save/record/already helpers."""

    def test_load_empty_history(self, tmp_path: Path):
        history = load_history(str(tmp_path))
        assert history == {}

    def test_save_and_load_round_trip(self, tmp_path: Path):
        data = {"file.txt": {"status": "success", "date": "2026-01-01"}}
        save_history(data, str(tmp_path))
        loaded = load_history(str(tmp_path))
        assert loaded == data

    def test_record_success_marks_file(self, tmp_path: Path):
        history: dict = {}
        record_success(history, "DSNO_001.txt", str(tmp_path))
        assert history["DSNO_001.txt"]["status"] == "success"
        assert "date" in history["DSNO_001.txt"]

    def test_already_downloaded_success(self):
        history = {"f.txt": {"status": "success"}}
        assert already_downloaded(history, "f.txt") is True

    def test_already_downloaded_legacy_sucesso(self):
        history = {"f.txt": {"status": "sucesso"}}
        assert already_downloaded(history, "f.txt") is True

    def test_not_yet_downloaded(self):
        assert already_downloaded({}, "f.txt") is False

    def test_load_corrupted_history(self, tmp_path: Path):
        path = tmp_path / "historico_downloads.json"
        path.write_text("{invalid json", encoding="utf-8")
        with pytest.raises(ConfigurationError, match="Malformed"):
            load_history(str(tmp_path))

    def test_history_persists_to_disk(self, tmp_path: Path):
        history: dict = {}
        record_success(history, "a.txt", str(tmp_path))
        record_success(history, "b.txt", str(tmp_path))

        reloaded = load_history(str(tmp_path))
        assert "a.txt" in reloaded
        assert "b.txt" in reloaded


# ── read_files_from_excel ────────────────────────────────────────────


def _make_download_excel(
    path: Path,
    rows: list[tuple[str, str, str | None]],
) -> Path:
    """Create an Excel file with ARGUMENT2, CREATION_DATE, STATUS columns.

    Each row is ``(dsno_filename, date_string, status)``.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ARGUMENT2", "CREATION_DATE", "STATUS"])
    for dsno, date_str, status in rows:
        dt = datetime.strptime(date_str, "%m-%d-%Y %H:%M:%S")
        ws.append([dsno, dt, status])
    wb.save(path)
    return path


class TestReadFilesFromExcel:
    """Tests for read_files_from_excel."""

    def test_reads_all_files(self, tmp_path: Path):
        path = _make_download_excel(
            tmp_path / "dl.xlsx",
            [
                ("DSNO_001.txt", "01-15-2026 10:00:00", "Open"),
                ("DSNO_002.txt", "01-16-2026 11:00:00", "Open"),
            ],
        )
        files = read_files_from_excel(
            str(path), "ARGUMENT2", "CREATION_DATE", "STATUS", "", "", ""
        )
        assert len(files) == 2

    def test_filters_by_date_range(self, tmp_path: Path):
        path = _make_download_excel(
            tmp_path / "dl.xlsx",
            [
                ("DSNO_001.txt", "01-15-2026 10:00:00", "Open"),
                ("DSNO_002.txt", "02-15-2026 10:00:00", "Open"),
            ],
        )
        files = read_files_from_excel(
            str(path),
            "ARGUMENT2",
            "CREATION_DATE",
            "STATUS",
            "01/01/2026 00:00:00",
            "31/01/2026 23:59:59",
            "",
        )
        assert files == ["DSNO_001.txt"]

    def test_filters_by_status(self, tmp_path: Path):
        path = _make_download_excel(
            tmp_path / "dl.xlsx",
            [
                ("DSNO_001.txt", "01-15-2026 10:00:00", "Open"),
                ("DSNO_002.txt", "01-16-2026 10:00:00", "Processed"),
            ],
        )
        files = read_files_from_excel(
            str(path),
            "ARGUMENT2",
            "CREATION_DATE",
            "STATUS",
            "",
            "",
            "Processed",
        )
        assert files == ["DSNO_002.txt"]

    def test_missing_dsno_column(self, tmp_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["OTHER_COL"])
        ws.append(["val"])
        path = tmp_path / "dl.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="not found"):
            read_files_from_excel(
                str(path), "ARGUMENT2", "CREATION_DATE", "STATUS", "", "", ""
            )

    def test_missing_date_column(self, tmp_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ARGUMENT2"])
        ws.append(["DSNO_001.txt"])
        path = tmp_path / "dl.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="not found"):
            read_files_from_excel(
                str(path), "ARGUMENT2", "CREATION_DATE", "STATUS", "", "", ""
            )
