"""Tests for dsno_processor.status_updater."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from dsno_processor.status_updater import (
    update_control_sheet_status,
    update_excel_status_for_downloaded,
)


# ── Helpers ──────────────────────────────────────────────────────────


def _create_control_excel(
    path: Path,
    rows: list[tuple[str, str | None]],
) -> Path:
    """Create a minimal control sheet Excel file.

    Each row is ``(dsno_filename, current_status)``.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ARGUMENT2", "STATUS"])
    for dsno, status in rows:
        ws.append([dsno, status])
    wb.save(path)
    return path


def _create_processed_dir(base: Path, filenames: list[str]) -> Path:
    """Create a Processed directory containing the given dummy files."""
    processed = base / "Processed"
    processed.mkdir(parents=True, exist_ok=True)
    for fn in filenames:
        (processed / fn).write_text("dummy", encoding="utf-8")
    return processed


# ── Tests ────────────────────────────────────────────────────────────


class TestUpdateControlSheetStatus:
    """Tests for update_control_sheet_status."""

    def test_updates_matching_rows(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [
                ("DSNO_001.txt", "Open"),
                ("DSNO_002.txt", "Open"),
                ("DSNO_003.txt", "Open"),
            ],
        )
        processed = _create_processed_dir(tmp_path, ["DSNO_001.txt", "DSNO_003.txt"])

        count = update_control_sheet_status(ctrl, processed)
        assert count == 2

        # Verify in the Excel file
        wb = openpyxl.load_workbook(ctrl)
        ws = wb.active
        statuses = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, 5)}
        assert statuses["DSNO_001.txt"] == "Processed"
        assert statuses["DSNO_002.txt"] == "Open"
        assert statuses["DSNO_003.txt"] == "Processed"

    def test_already_processed_not_counted(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [("DSNO_001.txt", "Processed")],
        )
        processed = _create_processed_dir(tmp_path, ["DSNO_001.txt"])
        count = update_control_sheet_status(ctrl, processed)
        assert count == 0

    def test_empty_processed_dir(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [("DSNO_001.txt", "Open")],
        )
        processed = _create_processed_dir(tmp_path, [])
        count = update_control_sheet_status(ctrl, processed)
        assert count == 0

    def test_control_sheet_not_found(self, tmp_path: Path):
        processed = _create_processed_dir(tmp_path, ["DSNO_001.txt"])
        count = update_control_sheet_status(tmp_path / "nope.xlsx", processed)
        assert count == 0

    def test_processed_dir_not_found(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [("DSNO_001.txt", "Open")],
        )
        count = update_control_sheet_status(ctrl, tmp_path / "missing")
        assert count == 0

    def test_creates_backup(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [("DSNO_001.txt", "Open")],
        )
        processed = _create_processed_dir(tmp_path, ["DSNO_001.txt"])
        update_control_sheet_status(ctrl, processed)

        backup_dir = tmp_path / "control_sheet_backups"
        assert backup_dir.exists()
        backups = list(backup_dir.glob("ctrl_*.xlsx"))
        assert len(backups) == 1

    def test_adds_status_column_if_missing(self, tmp_path: Path):
        """If the STATUS column doesn't exist, it should be created."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ARGUMENT2"])
        ws.append(["DSNO_001.txt"])
        ctrl = tmp_path / "ctrl.xlsx"
        wb.save(ctrl)

        processed = _create_processed_dir(tmp_path, ["DSNO_001.txt"])
        count = update_control_sheet_status(ctrl, processed)
        assert count == 1

        wb2 = openpyxl.load_workbook(ctrl)
        ws2 = wb2.active
        # STATUS column should have been added
        headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
        assert "STATUS" in headers


class TestUpdateDownloadedStatus:
    """Tests for update_excel_status_for_downloaded."""

    def test_updates_downloaded_status(self, tmp_path: Path):
        ctrl = _create_control_excel(
            tmp_path / "ctrl.xlsx",
            [
                ("DSNO_001.txt", "Open"),
                ("DSNO_002.txt", "Open"),
                ("DSNO_003.txt", "Open"),
            ],
        )

        # Create download directory and put matching files there
        dl_dir = tmp_path / "Downloads"
        dl_dir.mkdir()
        (dl_dir / "DSNO_001.txt").write_text("dummy")
        (dl_dir / "DSNO_003.txt").write_text("dummy")

        count = update_excel_status_for_downloaded(ctrl, dl_dir)
        assert count == 2

        # Verify status in excel
        wb = openpyxl.load_workbook(ctrl)
        ws = wb.active
        statuses = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, 5)}
        assert statuses["DSNO_001.txt"] == "Downloaded"
        assert statuses["DSNO_002.txt"] == "Open"
        assert statuses["DSNO_003.txt"] == "Downloaded"

    def test_sheet_not_found(self, tmp_path: Path):
        count = update_excel_status_for_downloaded(tmp_path / "nope.xlsx", tmp_path / "Downloads")
        assert count == 0

    def test_dl_dir_not_found(self, tmp_path: Path):
        ctrl = _create_control_excel(tmp_path / "ctrl.xlsx", [("DSNO_001.txt", "Open")])
        count = update_excel_status_for_downloaded(ctrl, tmp_path / "nope_dl")
        assert count == 0

