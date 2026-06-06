"""Tests for dsno_processor.database."""

from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

import pytest

from dsno_processor.database import (
    ControlRecord,
    ShipmentRecord,
    get_connection,
    get_db_path,
    init_db,
    insert_control_record,
    upsert_control_record,
    get_control_pairs,
    get_status_options,
    update_status,
    update_statuses_for_processed,
    insert_shipment,
    upsert_shipment,
    get_shipment_info,
    import_customer_sheet,
    import_control_sheet,
)


# ── Fixtures ─────────────────────────────────────────────────────────


@pytest.fixture
def db(tmp_path: Path):
    """Provide a fresh in-memory database with schema initialized."""
    conn = sqlite3.connect(":memory:", detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    init_db(conn)
    yield conn
    conn.close()


@pytest.fixture
def populated_db(db):
    """Database with sample control and shipment records."""
    records = [
        ControlRecord(
            invoice=1001,
            dsno_filename="DSNO001.txt",
            creation_date=datetime(2026, 3, 10),
            status="Open",
            freight_oracle="MARITIMA",
            freight_softway="MARITIMA",
        ),
        ControlRecord(
            invoice=1002,
            dsno_filename="DSNO002.txt",
            creation_date=datetime(2026, 3, 15),
            status="Processed",
            freight_oracle="AEREA",
            freight_softway=None,
        ),
        ControlRecord(
            invoice=1003,
            dsno_filename="DSNO003.txt",
            creation_date=datetime(2026, 4, 1),
            status=None,
            freight_oracle="MARITIMA",
            freight_softway="MARITIMA",
        ),
    ]
    for r in records:
        insert_control_record(db, r)

    shipments = [
        ShipmentRecord(invoice=1001, booking="BK-001", container="CNT-001"),
        ShipmentRecord(invoice=1002, booking="BK-002", container="CNT-002"),
    ]
    for s in shipments:
        insert_shipment(db, s)

    return db


# ── Schema / lifecycle ───────────────────────────────────────────────


class TestLifecycle:
    def test_get_db_path_default(self):
        path = get_db_path()
        assert path.name == "dsno_processor.db"

    def test_get_db_path_custom(self, tmp_path):
        path = get_db_path(tmp_path)
        assert path.parent == tmp_path
        assert path.name == "dsno_processor.db"

    def test_get_connection_creates_file(self, tmp_path):
        db_path = tmp_path / "test.db"
        conn = get_connection(db_path)
        assert db_path.exists()
        conn.close()

    def test_init_db_creates_tables(self, db):
        tables = db.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
        names = {r["name"] for r in tables}
        assert "tb_control" in names
        assert "tb_shipment_info" in names

    def test_init_db_idempotent(self, db):
        """Calling init_db twice does not raise."""
        init_db(db)
        init_db(db)


# ── tb_control CRUD ──────────────────────────────────────────────────


class TestControlCRUD:
    def test_insert_and_retrieve(self, db):
        rec = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 15),
            status="Open",
            freight_oracle="SEA",
        )
        row_id = insert_control_record(db, rec)
        assert row_id > 0

        row = db.execute("SELECT * FROM tb_control WHERE ID = ?", (row_id,)).fetchone()
        assert row["INVOICE"] == 100
        assert row["DSNO_FILENAME"] == "DSNO100.txt"
        assert row["STATUS"] == "Open"

    def test_unique_dsno_filename(self, db):
        rec = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 1),
        )
        insert_control_record(db, rec)
        with pytest.raises(sqlite3.IntegrityError):
            insert_control_record(db, rec)

    def test_upsert_updates_existing(self, db):
        rec = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 1),
            status="Open",
        )
        insert_control_record(db, rec)

        updated = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 2),
            status="Processed",
        )
        upsert_control_record(db, updated)

        row = db.execute(
            "SELECT * FROM tb_control WHERE DSNO_FILENAME = 'DSNO100.txt'"
        ).fetchone()
        assert row["STATUS"] == "Processed"

    def test_upsert_preserves_null_status(self, db):
        rec = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 1),
            status="Open",
        )
        insert_control_record(db, rec)

        updated = ControlRecord(
            invoice=100,
            dsno_filename="DSNO100.txt",
            creation_date=datetime(2026, 1, 1),
            status=None,
        )
        upsert_control_record(db, updated)

        row = db.execute(
            "SELECT STATUS FROM tb_control WHERE DSNO_FILENAME = 'DSNO100.txt'"
        ).fetchone()
        assert row["STATUS"] == "Open"  # preserved, not overwritten with None


# ── get_control_pairs ────────────────────────────────────────────────


class TestGetControlPairs:
    def test_filters_by_date_range(self, populated_db):
        pairs = get_control_pairs(
            populated_db,
            start=datetime(2026, 3, 1),
            end=datetime(2026, 3, 31),
        )
        assert len(pairs) == 2
        invoices = [p[0] for p in pairs]
        assert 1001 in invoices
        assert 1002 in invoices

    def test_filters_by_status(self, populated_db):
        pairs = get_control_pairs(
            populated_db,
            start=datetime(2026, 1, 1),
            end=datetime(2026, 12, 31),
            status_filter=["Open"],
        )
        # Should include: Open (1001) + NULL status (1003)
        invoices = [p[0] for p in pairs]
        assert 1001 in invoices
        assert 1003 in invoices
        assert 1002 not in invoices  # "Processed" not in filter

    def test_no_filter_returns_all(self, populated_db):
        pairs = get_control_pairs(
            populated_db,
            start=datetime(2026, 1, 1),
            end=datetime(2026, 12, 31),
        )
        assert len(pairs) == 3

    def test_returns_four_element_tuples(self, populated_db):
        pairs = get_control_pairs(
            populated_db,
            start=datetime(2026, 1, 1),
            end=datetime(2026, 12, 31),
        )
        for invoice, dsno, oracle, softway in pairs:
            assert isinstance(invoice, int)
            assert isinstance(dsno, str)


# ── get_status_options ───────────────────────────────────────────────


class TestGetStatusOptions:
    def test_returns_sorted_unique(self, populated_db):
        options = get_status_options(populated_db)
        assert options == ["Open", "Processed"]

    def test_returns_empty_if_less_than_two(self, db):
        insert_control_record(
            db,
            ControlRecord(
                invoice=1,
                dsno_filename="D1.txt",
                creation_date=datetime(2026, 1, 1),
                status="Open",
            ),
        )
        assert get_status_options(db) == []


# ── update_status ────────────────────────────────────────────────────


class TestUpdateStatus:
    def test_updates_existing(self, populated_db):
        result = update_status(populated_db, "DSNO001.txt", "Processed")
        assert result is True

        row = populated_db.execute(
            "SELECT STATUS FROM tb_control WHERE DSNO_FILENAME = 'DSNO001.txt'"
        ).fetchone()
        assert row["STATUS"] == "Processed"

    def test_returns_false_for_missing(self, populated_db):
        result = update_status(populated_db, "NONEXISTENT.txt", "Processed")
        assert result is False


class TestUpdateStatusesForProcessed:
    def test_bulk_update(self, populated_db):
        count = update_statuses_for_processed(
            populated_db, {"DSNO001.txt", "DSNO003.txt"}
        )
        assert count == 2

    def test_skips_already_processed(self, populated_db):
        count = update_statuses_for_processed(
            populated_db,
            {"DSNO002.txt"},  # already "Processed"
        )
        assert count == 0

    def test_empty_set(self, populated_db):
        assert update_statuses_for_processed(populated_db, set()) == 0


# ── tb_shipment_info CRUD ────────────────────────────────────────────


class TestShipmentCRUD:
    def test_insert_and_retrieve(self, db):
        rec = ShipmentRecord(invoice=200, booking="BK-200", container="CNT-200")
        row_id = insert_shipment(db, rec)
        assert row_id > 0

        info = get_shipment_info(db, 200)
        assert info is not None
        assert info.invoice == 200
        assert info.booking == "BK-200"
        assert info.container == "CNT-200"

    def test_get_returns_none_when_not_found(self, db):
        assert get_shipment_info(db, 999) is None

    def test_upsert_with_esn(self, db):
        rec = ShipmentRecord(esn=12345, invoice=200, booking="BK-1", container="CNT-1")
        insert_shipment(db, rec)

        updated = ShipmentRecord(
            esn=12345, invoice=200, booking="BK-2", container="CNT-2"
        )
        upsert_shipment(db, updated)

        info = get_shipment_info(db, 200)
        assert info.booking == "BK-2"

    def test_multiple_null_esn_allowed(self, db):
        """NULL ESNs are distinct — multiple inserts should succeed."""
        insert_shipment(
            db, ShipmentRecord(invoice=300, booking="BK-A", container="C-A")
        )
        insert_shipment(
            db, ShipmentRecord(invoice=300, booking="BK-B", container="C-B")
        )

        rows = db.execute(
            "SELECT COUNT(*) as cnt FROM tb_shipment_info WHERE INVOICE = 300"
        ).fetchone()
        assert rows["cnt"] == 2

    def test_get_returns_first_match(self, db):
        insert_shipment(db, ShipmentRecord(invoice=300, booking="FIRST", container="C"))
        insert_shipment(
            db, ShipmentRecord(invoice=300, booking="SECOND", container="C")
        )

        info = get_shipment_info(db, 300)
        assert info.booking == "FIRST"


# ── import_customer_sheet ────────────────────────────────────────────


class TestImportCustomerSheet:
    def test_import_from_excel(self, db, tmp_path):
        """Test import using a real Excel file created with openpyxl."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Invoice", "Booking/HAWB", "Container"])
        ws.append([5001, "BK-5001", "CNT-5001"])
        ws.append([5002, "BK-5002", "CNT-5002"])
        ws.append([None, "BK-SKIP", "CNT-SKIP"])  # no invoice → skip

        xlsx = tmp_path / "test_customer.xlsx"
        wb.save(xlsx)

        imported, skipped = import_customer_sheet(db, xlsx)
        assert imported == 2
        assert skipped == 1

        info = get_shipment_info(db, 5001)
        assert info is not None
        assert info.booking == "BK-5001"

    def test_import_missing_file(self, db):
        with pytest.raises(FileNotFoundError):
            import_customer_sheet(db, Path("/nonexistent.xlsx"))

    def test_import_missing_columns(self, db, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["WrongCol1", "WrongCol2"])
        ws.append([1, 2])
        xlsx = tmp_path / "bad.xlsx"
        wb.save(xlsx)

        with pytest.raises(ValueError, match="Missing required columns"):
            import_customer_sheet(db, xlsx)

    def test_import_idempotent_with_esn(self, db, tmp_path):
        """Re-importing the same data should not create duplicates when ESN is set."""
        rec = ShipmentRecord(esn=9999, invoice=7000, booking="BK", container="CNT")
        insert_shipment(db, rec)

        # Upserting the same ESN should update, not duplicate
        upsert_shipment(
            db,
            ShipmentRecord(esn=9999, invoice=7000, booking="BK-NEW", container="CNT"),
        )

        rows = db.execute(
            "SELECT COUNT(*) as cnt FROM tb_shipment_info WHERE ESN = 9999"
        ).fetchone()
        assert rows["cnt"] == 1

        info = get_shipment_info(db, 7000)
        assert info.booking == "BK-NEW"


# ── import_control_sheet ─────────────────────────────────────────────


class TestImportControlSheet:
    def test_import_control_from_excel(self, db, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "INVOICE",
            "ARGUMENT2",
            "CREATION_DATE",
            "STATUS",
            "FREIGHT_ORACLE",
            "FREIGHT_SOFTWAY",
            "Obs",
        ])
        ws.append([6001, "DSNO_6001.txt", "2026-05-01 10:00:00", "Open", "MARITIMA", "MARITIMA", "Note 1"])
        ws.append([6002, "DSNO_6002.txt", "2026-05-02 11:00:00", "Processed", "AEREA", "AEREA", "Note 2"])
        ws.append([None, "DSNO_SKIP.txt", "2026-05-03 12:00:00", "Open", "", "", ""])  # no invoice -> skip
        ws.append([6003, None, "2026-05-04 13:00:00", "Open", "", "", ""])  # no dsno -> skip

        xlsx = tmp_path / "test_control.xlsx"
        wb.save(xlsx)

        imported, updated, skipped = import_control_sheet(db, xlsx)
        assert imported == 2
        assert updated == 0
        assert skipped == 2

        # Check records in db
        res = db.execute("SELECT * FROM tb_control WHERE INVOICE = 6001").fetchone()
        assert res is not None
        assert res["DSNO_FILENAME"] == "DSNO_6001.txt"
        assert res["STATUS"] == "Open"
        assert res["FREIGHT_ORACLE"] == "MARITIMA"
        assert res["DESCRIPTION"] == "Note 1"

        # Now test updating
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "INVOICE",
            "ARGUMENT2",
            "CREATION_DATE",
            "STATUS",
            "FREIGHT_ORACLE",
            "FREIGHT_SOFTWAY",
            "Obs",
        ])
        ws.append([6001, "DSNO_6001.txt", "2026-05-01 10:00:00", "Processed", "MARITIMA", "MARITIMA", "Updated Note"])

        wb.save(xlsx)
        imported, updated, skipped = import_control_sheet(db, xlsx)
        assert imported == 0
        assert updated == 1
        assert skipped == 0

        res = db.execute("SELECT * FROM tb_control WHERE INVOICE = 6001").fetchone()
        assert res["STATUS"] == "Processed"
        assert res["DESCRIPTION"] == "Updated Note"

    def test_import_control_missing_file(self, db):
        with pytest.raises(FileNotFoundError):
            import_control_sheet(db, Path("/nonexistent_control.xlsx"))

    def test_import_control_missing_columns(self, db, tmp_path):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["WrongCol1", "WrongCol2"])
        ws.append([1, 2])
        xlsx = tmp_path / "bad_control.xlsx"
        wb.save(xlsx)

        with pytest.raises(ValueError, match="Missing required columns"):
            import_control_sheet(db, xlsx)
