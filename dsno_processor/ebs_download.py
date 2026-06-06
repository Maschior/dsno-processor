"""EBS File Download Automation — core module.

Provides reusable, parameterised functions for automating file downloads
from Oracle EBS, driven from the GUI.
"""

from __future__ import annotations

import json
import logging
import os
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException

from .exceptions import ConfigurationError, CanceledError, LoginError
from .i18n import t

logger = logging.getLogger(__name__)


# ── Config dataclass ─────────────────────────────────────────────────


@dataclass
class DownloadConfig:
    """All parameters needed for an EBS download run."""

    ebs_url: str
    customer_sheet_path: str
    download_dir: str
    email: str = ""
    password: str = ""
    dsno_col: str = "ARGUMENT2"
    date_col: str = "CREATION_DATE"
    status_col: str = "STATUS"
    date_start: str = ""
    date_end: str = ""
    status_filter: str = ""
    headless: bool = False
    folder_indices: list[int] = field(default_factory=lambda: [92, 95, 101])


# ── History tracking ─────────────────────────────────────────────────


def _history_path(download_dir: str) -> Path:
    return Path(download_dir) / "historico_downloads.json"


def load_history(download_dir: str) -> dict:
    """Load the download history from disk."""
    path = _history_path(download_dir)
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            logger.error("Malformed or empty history file: %s", e)
            raise ConfigurationError(
                "Malformed or corrupted history file.\n"
                "Delete the file or fix it manually.\n"
                f"File location: {path}"
            )
    return {}


def save_history(history: dict, download_dir: str) -> None:
    """Persist the download history to disk."""
    path = _history_path(download_dir)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def record_success(history: dict, filename: str, download_dir: str) -> None:
    """Mark a file as successfully downloaded in the history."""
    history[filename] = {
        "status": "success",
        "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    save_history(history, download_dir)


def already_downloaded(history: dict, filename: str) -> bool:
    """Check whether a file has already been downloaded."""
    entry = history.get(filename, {})
    return entry.get("status") in ("success", "sucesso")


# ── Excel ────────────────────────────────────────────────────────────


def read_files_from_excel(
    path: str,
    dsno_col: str,
    date_col: str,
    status_col: str,
    date_start: str,
    date_end: str,
    status_filter: str,
) -> list[str]:
    """Read file names from the customer spreadsheet, filtered by date range and status."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if dsno_col not in headers:
        raise ValueError(f"Column '{dsno_col}' not found. Available columns: {headers}")

    dsno_col_idx = headers.index(dsno_col)
    date_col_idx = headers.index(date_col) if date_col in headers else None
    status_col_idx = headers.index(status_col) if status_col in headers else None

    if date_col_idx is None:
        raise ValueError(f"Column '{date_col}' not found.")

    start = datetime.strptime(date_start, "%d/%m/%Y %H:%M:%S") if date_start else None
    end = datetime.strptime(date_end, "%d/%m/%Y %H:%M:%S") if date_end else None

    files: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[dsno_col_idx]
        date = row[date_col_idx]

        if status_col_idx is not None:
            status = (
                str(row[status_col_idx]) if row[status_col_idx] is not None else None
            )
        else:
            status = None

        # Parse the date
        if date is not None and not isinstance(date, datetime):
            if isinstance(date, str):
                date = datetime.strptime(date, "%m-%d-%Y %H:%M:%S")
            elif isinstance(date, (int, float)):
                date = datetime.fromtimestamp(date)
            else:
                continue

        if not value or date is None:
            continue

        # Apply date filter
        if start and end:
            if not (start <= date <= end):
                continue

        # Apply status filter
        if status_filter:
            if (
                status is None
                or status.lower().strip() != status_filter.lower().strip()
            ):
                continue

        files.append(str(value).strip())

    logger.info("%d file(s) found in spreadsheet.", len(files))
    return files


# ── Browser ──────────────────────────────────────────────────────────


def start_browser(download_dir: str, headless: bool = False) -> webdriver.Chrome:
    """Start a Chrome browser configured for file downloads."""
    options = webdriver.ChromeOptions()
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
    }
    options.add_experimental_option("prefs", prefs)
    # `detach` is removed to prevent zombie processes if the app closes unexpectedly
    options.add_argument("--safebrowsing-disable-download-protection")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-features=SafeBrowsingEnhancedProtection")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        logger.info("Headless mode enabled — browser running in background.")

    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {
            "behavior": "allow",
            "downloadPath": download_dir,
        },
    )

    return driver


def open_url(driver: webdriver.Chrome, url: str) -> None:
    """Navigate the browser to the given URL."""
    driver.get(url)
    logger.info("Browser opened.")


def _stoppable_sleep(seconds: float, cancel_event) -> None:
    if not cancel_event:
        time.sleep(seconds)
        return
    end_time = time.time() + seconds
    while time.time() < end_time:
        if cancel_event.is_set():
            raise CanceledError("Cancelled by user")
        time.sleep(0.1)


def perform_microsoft_login(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    email: str,
    password: str,
    target_url: str = "",
    cancel_event=None,
) -> None:
    """Automate Microsoft SSO login: click sign-in, enter email, enter password."""
    logger.info("Starting automatic Microsoft login...")

    if cancel_event and cancel_event.is_set():
        raise CanceledError("Cancelled by user")

    # Step 1: Click on sign-in button
    try:
        sign_in_btn = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".middle.ext-middle"))
        )
        sign_in_btn.click()
        logger.info("Clicked sign-in button.")
        _stoppable_sleep(3, cancel_event)
    except CanceledError:
        raise
    except Exception:
        logger.info(
            "Sign-in button not found — possibly already on login page or page down."
        )

    # Check if page is down
    if (
        "err_connection" in driver.page_source.lower()
        or "this site can't be reached" in driver.page_source.lower()
    ):
        raise Exception("Login screen or EBS appears to be down (connection error).")

    if cancel_event and cancel_event.is_set():
        raise CanceledError("Cancelled by user")

    # Step 2: Enter email
    try:
        email_input = wait.until(EC.element_to_be_clickable((By.ID, "i0116")))
        email_input.clear()
        email_input.send_keys(email)
        logger.info("Email entered.")
        _stoppable_sleep(1, cancel_event)
        email_input.send_keys(Keys.RETURN)
        _stoppable_sleep(3, cancel_event)
    except CanceledError:
        raise
    except Exception as e:
        logger.warning("Error entering email, or login screen not available: %s", e)
        raise Exception(
            "Failed to find or interact with Email input. Is the login screen loaded?"
        )

    if cancel_event and cancel_event.is_set():
        raise CanceledError("Cancelled by user")

    # Step 3: Enter password
    try:
        password_input = wait.until(EC.element_to_be_clickable((By.ID, "i0118")))
        password_input.clear()
        password_input.send_keys(password)
        logger.info("Password entered.")
        _stoppable_sleep(1, cancel_event)
        password_input.send_keys(Keys.RETURN)

        # Check for incorrect password
        for _ in range(10):
            if cancel_event and cancel_event.is_set():
                raise CanceledError("Cancelled by user")

            # Use find_elements to avoid NoSuchElementException if not found
            err_elements = driver.find_elements(By.ID, "passwordError")
            if err_elements and err_elements[0].is_displayed():
                err_msg = err_elements[0].text
                raise LoginError(f"Login failed: {err_msg}")

            # If we see signs of redirecting, we can break early
            if (
                "kmsi" in driver.current_url.lower()
                or "ebs" in driver.current_url.lower()
            ):
                break

            _stoppable_sleep(0.5, cancel_event)

    except LoginError as e:
        logger.error(f"Login failed: {str(e)}")
        raise
    except CanceledError as e:
        logger.error(f"Operation cancelled: {str(e)}")
        raise
    except Exception as e:
        if "Login failed" in str(e) or "Cancelled" in str(e):
            raise
        logger.warning("Error entering password: %s", e)
        raise

    logger.info("Waiting for EBS redirect...")
    try:
        xxdba_menu = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//div[normalize-space()='XXDBA Utilities']")
            )
        )
        xxdba_menu.click()
        logger.info("Clicked 'XXDBA Utilities' menu.")
        _stoppable_sleep(3, cancel_event)
    except CanceledError:
        raise
    except Exception:
        logger.info(
            "'XXDBA Utilities' menu not found "
            "(may already be on the correct page or layout differs)."
        )

    if cancel_event and cancel_event.is_set():
        raise CanceledError("Cancelled by user")

    if target_url:
        logger.info("Reloading target URL to ensure correct page...")
        driver.get(target_url)
        _stoppable_sleep(5, cancel_event)
        # Final check if EBS is down after redirect
        if (
            "err_connection" in driver.page_source.lower()
            or "this site can't be reached" in driver.page_source.lower()
        ):
            raise Exception("EBS appears to be down (connection error) after login.")

    logger.info("Automatic login complete.")


def list_folders(driver: webdriver.Chrome, wait: WebDriverWait) -> list[str]:
    """List available folders in the EBS file path dropdown."""
    try:
        select_el = wait.until(EC.presence_of_element_located((By.ID, "FilePath")))
        select = Select(select_el)
        folders = []
        for i, option in enumerate(select.options):
            text = option.text.strip() or "(empty)"
            folders.append(f"[{i}] {text}")
        return folders
    except CanceledError:
        raise
    except Exception as e:
        logger.warning("Could not list folders: %s", e)
        return []


# ── Download ─────────────────────────────────────────────────────────


def _select_folder(driver, wait, index, cancel_event=None):
    """Select a folder by index in the EBS dropdown."""
    select_el = wait.until(EC.presence_of_element_located((By.ID, "FilePath")))
    select = Select(select_el)
    select.select_by_index(index)
    _stoppable_sleep(3, cancel_event)


def _file_found(driver):
    """Check if the file name input field is available."""
    try:
        field = driver.find_element(By.ID, "FileName")
        return field.is_displayed() and field.is_enabled()
    except Exception:
        return False


def _download_via_requests(
    driver: webdriver.Chrome, filename: str, download_dir: str
) -> bool:
    """Download the file using requests by reusing Selenium's cookies and form state."""
    try:
        session = requests.Session()
        # 1. Transfer cookies
        for cookie in driver.get_cookies():
            session.cookies.set(cookie["name"], cookie["value"])

        # 2. Transfer User-Agent to look like the same browser
        user_agent = driver.execute_script("return navigator.userAgent")
        session.headers.update({"User-Agent": user_agent})

        # 3. Extract form data from the current page using JavaScript
        # This is atomic on the browser side and avoids StaleElementReferenceException
        # that often happens when scraping dynamic Oracle EBS forms manually.
        data = driver.execute_script("""
            var form = document.forms['DefaultFormName'];
            if (!form) return null;
            var data = {};
            for (var i = 0; i < form.elements.length; i++) {
                var el = form.elements[i];
                if (el.name) {
                    data[el.name] = el.value;
                }
            }
            return data;
        """)

        if not data:
            logger.warning("     Form 'DefaultFormName' not found on current page.")
            return False

        # 5. Inject/Override parameters for the Download event
        # This simulates the JS call: submitForm('DefaultFormName',1,{event:'Download',source:'Download'});
        data["event"] = "Download"
        data["source"] = "Download"
        # Ensure the filename is correct even if Selenium send_keys just happened
        data["FileName"] = filename

        url = driver.current_url
        logger.info("     Initiating POST request via requests session...")

        response = session.post(url, data=data, stream=True, timeout=60)

        # 6. Validate response
        content_type = response.headers.get("Content-Type", "").lower()
        if "text/html" in content_type:
            # Check if it's a small error page
            if len(response.content) < 5000:
                logger.warning(
                    "     Server returned HTML instead of a file. Likely an error."
                )
                return False

        # If we got here, we likely have the file or a very large HTML (unlikely)
        # Check Content-Disposition for filename or just save it
        disp = response.headers.get("Content-Disposition", "")
        save_name = filename
        if "filename=" in disp:
            save_name = disp.split("filename=")[-1].strip('"').strip("'")

        # Ensure it starts with DSNO if that's the rule
        if not save_name.upper().startswith("DSNO") and "text/html" in content_type:
            logger.warning("     Incorrect file type detected in requests response.")
            return False

        file_path = os.path.join(download_dir, save_name)
        with open(file_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        # Post-download verification
        if os.path.getsize(file_path) == 0:
            logger.warning(
                "     Downloaded file is empty (0 bytes). Deleting: %s", save_name
            )
            os.remove(file_path)
            return False

        logger.info("     Download successful via requests: %s", save_name)
        return True

    except CanceledError:
        raise
    except Exception as e:
        logger.error("     Unexpected error in requests download: %s", e)
        return False


def _attempt_download(
    driver, wait, filename, config: DownloadConfig, cancel_event=None
):
    """Try to download a file from the current folder."""
    try:
        # Wrap the primary input interaction in a retry loop to handle
        # StaleElementReferenceException, which often happens in EBS
        # when the page is still "settling" after a folder selection.
        field = None
        for attempt in range(3):
            if cancel_event and cancel_event.is_set():
                raise CanceledError("Cancelled by user")
            try:
                field = wait.until(EC.element_to_be_clickable((By.ID, "FileName")))
                field.clear()
                field.send_keys(filename)
                # Confirm text was actually entered
                if field.get_attribute("value") == filename:
                    break
            except StaleElementReferenceException:
                if attempt == 2:
                    raise
                _stoppable_sleep(1, cancel_event)
                continue

        _stoppable_sleep(5, cancel_event)

        # TAB might also need protection
        for attempt in range(3):
            if cancel_event and cancel_event.is_set():
                raise CanceledError("Cancelled by user")
            try:
                field = driver.find_element(By.ID, "FileName")
                field.send_keys(Keys.TAB)
                break
            except StaleElementReferenceException:
                if attempt == 2:
                    raise
                _stoppable_sleep(1, cancel_event)
                continue

        _stoppable_sleep(2, cancel_event)

        # 1. Check for EBS LOV/Popup (indicates file not found uniquely)
        popup_identifiers = [
            (By.ID, "lovPopUp_FileName"),
            (By.ID, "closeAnchorlovPopUp_FileName"),
            (By.ID, "FNDDIALOG"),
            (By.CSS_SELECTOR, "div[id*='lovPopUp']"),
            (By.CLASS_NAME, "xmb"),
        ]
        for by, sel in popup_identifiers:
            try:
                els = driver.find_elements(by, sel)
                if els and any(e.is_displayed() for e in els):
                    logger.info(
                        "EBS Popup/Dialog detected — file not found in this folder."
                    )
                    try:
                        driver.execute_script(
                            "document.getElementById('closeAnchorlovPopUp_FileName')?.click();"
                        )
                    except Exception:
                        logger.debug("Unable to close EBS popup", exc_info=True)
                    return False
            except Exception as e:
                # If element becomes stale during checking, it likely means the page
                # is refreshing/changing, which is fine to ignore in this detection loop.
                if "stale" in str(e).lower():
                    continue
                pass

        # 3. Use requests to download (reusing cookies and form state)
        if _download_via_requests(driver, filename, config.download_dir):
            return True

        return False

    except CanceledError:
        raise
    except Exception as e:
        logger.warning("Error in _attempt_download: %s", e)
        return False


def _reset_form(driver, url, cancel_event=None):
    """Reset the EBS form by reloading the page."""
    driver.get(url)
    _stoppable_sleep(3, cancel_event)


def download_file(
    driver: webdriver.Chrome,
    wait: WebDriverWait,
    filename: str,
    config: DownloadConfig,
    cancel_event=None,
) -> bool:
    """Try to download a file, iterating through all configured folders."""
    for i, index in enumerate(config.folder_indices, 1):
        if cancel_event and cancel_event.is_set():
            return False
        logger.info(
            "Trying folder %d/%d (index %d)...",
            i,
            len(config.folder_indices),
            index,
        )
        try:
            _select_folder(driver, wait, index, cancel_event)
        except CanceledError:
            raise
        except Exception as e:
            if cancel_event and cancel_event.is_set():
                raise CanceledError("Cancelled by user")
            logger.warning("Could not select folder %d: %s", i, e)
            _reset_form(driver, config.ebs_url, cancel_event)
            continue

        if not _file_found(driver):
            logger.info("File input not available in this folder.")
            _reset_form(driver, config.ebs_url, cancel_event)
            continue

        if _attempt_download(driver, wait, filename, config, cancel_event):
            return True
        _reset_form(driver, config.ebs_url, cancel_event)

    return False


# ── Orchestrator ─────────────────────────────────────────────────────


def run_download(
    config: DownloadConfig, progress_callback=None, cancel_event=None
) -> dict:
    """Run the full download flow.

    Args:
        config: Download configuration.
        progress_callback: Optional ``(event, data_dict)`` callable for
            real-time progress updates consumed by the GUI dashboard.
        cancel_event: Optional threading.Event to abort the operation.
    Returns:
        A summary dict with ``success``, ``skipped``, ``failures`` keys.
    """

    def _cb(event: str, data: dict | None = None) -> None:
        if progress_callback:
            progress_callback(event, data or {})

    # 1. Load history
    _cb("phase", {"text": "Loading history..."})
    history = load_history(config.download_dir)
    already_done = [
        k for k, v in history.items() if v["status"] in ("success", "sucesso")
    ]
    if already_done:
        logger.info("History: %d file(s) already downloaded.", len(already_done))

    # 2. Read spreadsheet
    _cb("phase", {"text": "Reading spreadsheet..."})
    all_files = read_files_from_excel(
        config.customer_sheet_path,
        config.dsno_col,
        config.date_col,
        config.status_col,
        config.date_start,
        config.date_end,
        config.status_filter,
    )

    # 3. Filter already downloaded
    pending = [f for f in all_files if not already_downloaded(history, f)]
    skipped_list = [f for f in all_files if already_downloaded(history, f)]
    skipped_count = len(skipped_list)

    # Report total (including skipped) to dashboard
    _cb("total", {"count": len(all_files)})

    # Report skipped items individually
    for f in skipped_list:
        _cb("skipped", {"name": f, "detail": "Already downloaded"})

    if skipped_count:
        logger.info("%d file(s) skipped (already downloaded).", skipped_count)
    logger.info("%d file(s) to download.", len(pending))

    if not pending:
        logger.info("All files have already been downloaded!")
        _cb("finished", {})
        return {"success": 0, "skipped": skipped_count, "failures": []}

    # 4. Start browser
    _cb("phase", {"text": "Starting browser..."})
    os.makedirs(config.download_dir, exist_ok=True)
    driver = start_browser(config.download_dir, headless=config.headless)
    success_count = 0
    failure_list: list[str] = []

    try:
        wait = WebDriverWait(driver, 15)

        # 5. Open URL and auto-login
        _cb("phase", {"text": "Opening EBS..."})
        open_url(driver, config.ebs_url)
        if config.email and config.password:
            _cb("phase", {"text": "Performing automatic login..."})
            perform_microsoft_login(
                driver,
                wait,
                config.email,
                config.password,
                config.ebs_url,
                cancel_event,
            )

        folders = list_folders(driver, wait)
        for f in folders:
            logger.info("  %s", f)

        # 6. Download files
        logger.info("Starting downloads (%d file(s))...", len(pending))

        for i, filename in enumerate(pending, 1):
            if cancel_event and cancel_event.is_set():
                logger.info("Download process cancelled.")
                _cb("error", {"name": "System", "detail": "Cancelled by user"})
                raise CanceledError("Cancelled by user")

            _cb("phase", {"text": f"Downloading {filename} ({i}/{len(pending)})..."})
            logger.info("[%d/%d] %s", i, len(pending), filename)
            found = download_file(driver, wait, filename, config, cancel_event)
            if cancel_event and cancel_event.is_set():
                raise CanceledError("Cancelled by user")

            if found:
                record_success(history, filename, config.download_dir)
                logger.info("Download started!")
                success_count += 1
                _cb("success", {"name": filename, "detail": t("dash.downloaded")})
            else:
                logger.warning(
                    "Not found in any of the %d folders.", len(config.folder_indices)
                )
                failure_list.append(filename)
                _cb("error", {"name": filename, "detail": "Not found in folders"})
            time.sleep(1)

        # 7. Report
        logger.info("=" * 55)
        logger.info("  Success:  %d file(s)", success_count)
        logger.info("  Skipped:  %d file(s)", skipped_count)
        logger.info("  Failed:   %d file(s)", len(failure_list))
        if failure_list:
            for f in failure_list:
                logger.info("    - %s", f)
        logger.info("  Downloads: %s", config.download_dir)

    except LoginError as e:
        _cb("error", {"name": "Login", "detail": str(e)})
        raise
    except CanceledError as e:
        _cb("cancelled", {"name": "Process", "detail": str(e)})
        raise
    except Exception as e:
        _cb("error", {"name": "Process", "detail": str(e)})
        raise
    finally:
        _cb("finished", {})
        try:
            driver.quit()
        except Exception:
            logger.debug("Unable to close browser driver cleanly", exc_info=True)

    return {
        "success": success_count,
        "skipped": skipped_count,
        "failures": failure_list,
    }
