"""Update the STATUS column in the internal control sheet."""

from __future__ import annotations

import logging
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl

log = logging.getLogger(__name__)


def update_control_sheet_status(
    control_sheet_path: Path | str,
    processed_dir: Path | str,
) -> int:
    """Read the control sheet and update STATUS to 'Processed'.

    Checks if the DSNO (from the ARGUMENT2 column) is physically present
    in the `processed_dir`. If so, updates its STATUS column.

    Args:
        control_sheet_path: Path to the Excel control sheet.
        processed_dir: Path to the Processed directory.

    Returns:
        The number of rows updated.
    """
    cs_path = Path(control_sheet_path)
    pd_path = Path(processed_dir)

    if not cs_path.exists():
        log.error("Control sheet not found for status update at %s", cs_path)
        return 0

    if not pd_path.exists():
        log.error("Processed directory not found at %s", pd_path)
        return 0

    processed_files = {f.name for f in pd_path.iterdir() if f.is_file()}
    if not processed_files:
        log.info("No files in Processed directory. Skipping status update.")
        return 0

    log.info("Updating STATUS column in control sheet: %s", cs_path.name)

    try:
        backup_dir = cs_path.parent / "control_sheet_backups"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{cs_path.stem}_{timestamp}{cs_path.suffix}"
        backup_path = backup_dir / backup_filename
        shutil.copy2(cs_path, backup_path)
        log.info("Created backup of control sheet at %s", backup_path)
    except Exception as exc:
        log.error("Failed to create control sheet backup: %s", exc)
        return 0

    try:
        wb = openpyxl.load_workbook(cs_path)
        ws = wb.active

        header_row = 1
        dsno_col_idx = None
        status_col_idx = None

        # Find columns
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val is not None:
                col_name = str(cell_val).strip().upper()
                if col_name == "ARGUMENT2":
                    dsno_col_idx = col_idx
                elif col_name == "STATUS":
                    status_col_idx = col_idx

        if dsno_col_idx is None:
            log.error("Column 'ARGUMENT2' (DSNO) not found in control sheet.")
            return 0

        if status_col_idx is None:
            # Add STATUS column if missing
            status_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=status_col_idx, value="STATUS")
            log.info("Added 'STATUS' column to control sheet.")

        updates_made = 0
        for row_idx in range(2, ws.max_row + 1):
            dsno_val = ws.cell(row=row_idx, column=dsno_col_idx).value
            if dsno_val:
                dsno_name = str(dsno_val).strip()
                if dsno_name in processed_files:
                    current_status = ws.cell(row=row_idx, column=status_col_idx).value
                    if current_status != "Processed":
                        ws.cell(row=row_idx, column=status_col_idx, value="Processed")
                        updates_made += 1

        if updates_made > 0:
            wb.save(cs_path)
            log.info("Updated STATUS to 'Processed' for %d rows.", updates_made)
        else:
            log.info("No new rows required STATUS update.")

        return updates_made
    except PermissionError as exc:
        log.error("Permission error while updating control sheet: %s", exc)
        log.error("Please close the control sheet if it is open and try again.")
        return 0
    except Exception as exc:
        log.exception("Failed to update status in control sheet: %s", exc)
        return 0


def update_excel_status_for_downloaded(
    sheet_path: Path | str,
    download_dir: Path | str,
    dsno_col: str = "ARGUMENT2",
    status_col: str = "STATUS",
) -> int:
    """Read the spreadsheet and update STATUS to 'Downloaded' for files in download_dir.

    Checks if the DSNO (from the dsno_col column) is physically present
    in the `download_dir`. If so, updates its status_col column to 'Downloaded'.
    """
    s_path = Path(sheet_path)
    dd_path = Path(download_dir)

    if not s_path.exists():
        log.error("Spreadsheet not found for status update at %s", s_path)
        return 0

    if not dd_path.exists():
        log.error("Download directory not found at %s", dd_path)
        return 0

    downloaded_files = {f.name for f in dd_path.iterdir() if f.is_file()}
    if not downloaded_files:
        log.info("No files in download directory. Skipping status update.")
        return 0

    log.info("Updating STATUS column in spreadsheet: %s", s_path.name)

    try:
        backup_dir = s_path.parent / "control_sheet_backups"
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{s_path.stem}_download_backup_{timestamp}{s_path.suffix}"
        backup_path = backup_dir / backup_filename
        shutil.copy2(s_path, backup_path)
        log.info("Created backup of spreadsheet at %s", backup_path)
    except Exception as exc:
        log.error("Failed to create spreadsheet backup: %s", exc)
        return 0

    try:
        wb = openpyxl.load_workbook(s_path)
        ws = wb.active

        header_row = 1
        dsno_col_idx = None
        status_col_idx = None

        target_dsno_col = str(dsno_col).strip().upper()
        target_status_col = str(status_col).strip().upper()

        # Find columns
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col_idx).value
            if cell_val is not None:
                col_name = str(cell_val).strip().upper()
                if col_name == target_dsno_col:
                    dsno_col_idx = col_idx
                elif col_name == target_status_col:
                    status_col_idx = col_idx

        if dsno_col_idx is None:
            log.error("Column '%s' not found in spreadsheet.", dsno_col)
            return 0

        if status_col_idx is None:
            # Add status column if missing
            status_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=status_col_idx, value=status_col)
            log.info("Added '%s' column to spreadsheet.", status_col)

        updates_made = 0
        for row_idx in range(2, ws.max_row + 1):
            dsno_val = ws.cell(row=row_idx, column=dsno_col_idx).value
            if dsno_val:
                dsno_name = str(dsno_val).strip()
                if dsno_name in downloaded_files:
                    current_status = ws.cell(row=row_idx, column=status_col_idx).value
                    if current_status != "Downloaded":
                        ws.cell(
                            row=row_idx, column=status_col_idx, value="Downloaded"
                        )
                        updates_made += 1

        if updates_made > 0:
            wb.save(s_path)
            log.info("Updated status to 'Downloaded' for %d rows.", updates_made)
        else:
            log.info("No new rows required status update.")

        return updates_made
    except PermissionError as exc:
        log.error("Permission error while updating spreadsheet: %s", exc)
        log.error("Please close the spreadsheet if it is open and try again.")
        return 0
    except Exception as exc:
        log.exception("Failed to update status in spreadsheet: %s", exc)
        return 0

